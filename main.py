import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook


# Function to search for the data
def search_data():
    id_search = entry_id.get()
    name_search = entry_name.get()
    load_data(id_search, name_search)  # Pass the ID and Name search terms to load_data


# Function to clear all the entry widgets and reset the Treeview
def clear_entries():
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_usd.delete(0, tk.END)
    entry_riel.delete(0, tk.END)
    load_data()  # Call load_data without arguments to reset the Treeview


# Update the update_data function to handle adding new data
def update_data():
    try:
        id_search = entry_id.get()
        name_to_find = entry_name.get()  # Also get the Name 
        usd_value = entry_usd.get()
        riel_value = entry_riel.get()

        # Attempt to convert USD and Riel values to integers
        usd_value = int(usd_value) if usd_value else 0
        riel_value = int(riel_value) if riel_value else 0

        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value).zfill(3) == id_search.zfill(3):  # Match ID
                row[1].value = name_to_find  # Update or add the Name
                row[2].value = usd_value  # Update or add the USD value
                row[3].value = riel_value  # Update or add the Riel value
                found = True
                break  # Can break, since we found the specific row

        if not found:
            # Add a new row with the new data
            sheet.append([id_search.zfill(3), name_to_find, usd_value, riel_value])
            print("New data added successfully.")

        wb.save('.\data.xlsx')
        print("Data updated successfully.")
        refresh_data()
        update_totals()  # Update totals after adding or updating data
    except ValueError:
        print("Please enter a valid number for USD and Riel.")
    except Exception as e:
        print(f"An error occurred: {e}")
    



# Function to update the totals
def update_totals():
    total_guests = 0
    total_usd = 0
    total_riel = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] > 0 or row[3] > 0:  # Check if USD or Riel is greater than 0
            total_guests += 1
        total_usd += row[2] if row[2] else 0
        total_riel += row[3] if row[3] else 0
    label_total_guests.config(text=f"Total Guests: {total_guests}")
    label_total_usd.config(text=f"Total USD: {total_usd}")
    label_total_riel.config(text=f"Total Riel: {total_riel}")


# Function to fill the entry widgets with the data from the selected row
def on_tree_select(event):
    # Check if there is a selection
    if tree.selection():
        selected_item = tree.selection()[0]  # Get selected item
        # Get the values of the selected row
        selected_row = tree.item(selected_item, 'values')
        # Clear the entry fields
        clear_entries()
        # Insert the data into the entry fields
        entry_id.insert(0, selected_row[0])
        entry_name.insert(0, selected_row[1])
        entry_usd.insert(0, selected_row[2])
        entry_riel.insert(0, selected_row[3])
    else:
        print("No item selected")


# Function to load data from Excel into the Treeview with search functionality
def load_data(id_search=None, name_search=None):
    for item in tree.get_children():
        tree.delete(item)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the ID matches the search term, if provided
        id_match = str(row[0]).zfill(3) == id_search.zfill(3) if id_search else True
        # Check if the Name matches the search term, if provided and not None
        name_match = name_search.lower() in (row[1] or "").lower() if name_search else True
        # If both ID and Name match (or are not provided), insert the row into the Treeview
        if id_match and name_match:
            tree.insert("", tk.END, values=row)


# Function to refresh the Treeview with updated data
def refresh_data():
    for item in tree.get_children():
        tree.delete(item)
    load_data()

    
    
# Load the workbook and select the active sheet
wb = load_workbook('.\data.xlsx')
sheet = wb.active   


# Create the main window
root = tk.Tk()
root.title("Wedding Guest Book")


# Create the entry widgets
entry_id = tk.Entry(root, width=40)
entry_id.grid(row=0, column=1)
entry_name = tk.Entry(root, width=40)
entry_name.grid(row=1, column=1)
entry_usd = tk.Entry(root, width=40)
entry_usd.grid(row=2, column=1)
entry_riel = tk.Entry(root, width=40)
entry_riel.grid(row=3, column=1)

# Create the labels for totals
label_total_guests = tk.Label(root, text="Total Guests: 0")
label_total_guests.grid(row=5, column=0)
label_total_usd = tk.Label(root, text="Total USD: 0")
label_total_usd.grid(row=5, column=1)
label_total_riel = tk.Label(root, text="Total Riel: 0")
label_total_riel.grid(row=5, column=2)


# Create the label widgets
label_id = tk.Label(root, text="ID")
label_id.grid(row=0, column=0)
label_name = tk.Label(root, text="Name")
label_name.grid(row=1, column=0)
label_usd = tk.Label(root, text="USD")
label_usd.grid(row=2, column=0)
label_riel = tk.Label(root, text="Riel")
label_riel.grid(row=3, column=0)

# Create the button widgets
button_search = tk.Button(root, text="Search", command=search_data,width=20)
button_search.grid(row=4, column=0)
# Create the 'Clear' button and place it on the grid
button_clear = tk.Button(root, text="Clear", command=clear_entries,width=20)
button_clear.grid(row=4, column=2)
button_update = tk.Button(root, text="Update", command=update_data,width=20)
button_update.grid(row=4, column=1)

# Create the Treeview widget
tree = ttk.Treeview(root, columns=("ID", "Name", "USD", "Riel"), show='headings')
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("USD", text="USD")
tree.heading("Riel", text="Riel")

tree.grid(row=7, columnspan=5, sticky='nsew')
# Bind the select event of the tree to the on_tree_select function
tree.bind('<<TreeviewSelect>>', on_tree_select)
# Load data into the Treeview
load_data()

# Run the main loop

root.mainloop() #ok

